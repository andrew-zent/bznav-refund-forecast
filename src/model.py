"""
Phase 2 v2 코호트 분산 모델 — 학습 + 예측 + 검증.

입력: data/deals_raw.json (extract_pipedrive.py 출력)
출력: output/forecast.json, output/backtest.json
"""
import json
import sys
from datetime import datetime, timezone
from collections import defaultdict
from pathlib import Path

import numpy as np

from config import (
    FIELD_MAP_BY_NAME, PIPELINE_REGULAR, PIPELINE_COLLECTION,
    STATUS_EXCLUDE, CHAIN_DIST_MAX_OFF, ROLLING_WINDOW,
    APP_FALLBACK_WINDOW, COLLECTION_MA_WINDOW, SEASON_ADJUSTMENT,
    CORP_PIPELINE_REGULAR, CORP_PIPELINE_COLLECTION,
)

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ── helpers ──────────────────────────────────────────
def ym(d: datetime) -> int:
    return d.year * 12 + (d.month - 1)

def ym_label(m: int) -> str:
    return f"{m // 12:04d}-{m % 12 + 1:02d}"

def month_of(m: int) -> int:
    return m % 12 + 1

def parse_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def to_num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


# ── data loading ─────────────────────────────────────
def load_deals() -> list[dict]:
    """JSON 또는 xlsx에서 deals 로드 → 내부 포맷."""
    # 1. slim JSON (extract_pipedrive.py의 최적화 출력) 우선
    slim_path = DATA_DIR / "deals_slim.json"
    if slim_path.exists():
        return _load_slim_json(slim_path)
    # 2. raw JSON (legacy) fallback
    raw_path = DATA_DIR / "deals_raw.json"
    if raw_path.exists():
        return _load_json(raw_path)
    # 3. xlsx fallback
    xlsx_files = list(DATA_DIR.glob("*.xlsx"))
    if xlsx_files:
        return _load_xlsx(xlsx_files[0])
    print("ERROR: data/deals_slim.json, deals_raw.json, 또는 *.xlsx 필요", file=sys.stderr)
    sys.exit(1)


def _load_slim_json(path: Path) -> list[dict]:
    """slim JSON (이미 변수명 매핑 완료) 직접 로드."""
    print(f"  Loading slim JSON: {path.name} ({path.stat().st_size / 1e6:.1f} MB)")
    raw = json.loads(path.read_text())
    # status 매핑 (slim에서 이미 pipeline 이름은 매핑됨)
    status_map = {"open": "진행 중", "won": "성사됨", "lost": "실패"}
    for d in raw:
        s = d.get("status", "")
        d["status"] = status_map.get(s, s)
    print(f"      loaded {len(raw):,} deals")
    return raw


def _load_json(path: Path) -> list[dict]:
    """Pipedrive API JSON → 내부 포맷.
    필드 key는 deal_fields.json으로 매핑."""
    print(f"  Loading JSON: {path.name}")

    # deal_fields.json → hash key to our variable name
    fields_path = DATA_DIR / "deal_fields.json"
    key_to_var = {}
    pipeline_map = {}  # pipeline_id -> name
    if fields_path.exists():
        fields = json.loads(fields_path.read_text())
        # Build name_fragment → var mapping
        name_fragments = {
            "✔ 신청일자": "apply_date",
            "📍 결제금액-알림톡발송": "payment_amount",
            "✍ 결정 환급액-알림톡발송": "decision_amount",
            "✔ 조회 환급액": "apply_amount",
            "✔ 신고일자": "filing_date",
            "✍ 신고 환급액-알림톡발송": "filing_amount",
            "✍ 결정일자": "decision_date",
            "💸 결제일자": "payment_date",
            "감면only 여부": "is_only_gam",
        }
        for key, info in fields.items():
            # Skip currency/통화 fields — they contain 'KRW' not amounts
            if info.get("field_type") == "varchar" and ("통화" in info["name"] or key.endswith("_currency")):
                continue
            for frag, var_name in name_fragments.items():
                if frag in info["name"]:
                    key_to_var[key] = var_name
                    break
            # pipeline field options
            if info["name"] == "파이프라인" and info.get("options"):
                pipeline_map = info["options"]

    print(f"      mapped {len(key_to_var)} custom fields")

    # Pipedrive pipeline_id → name mapping via API response
    # (pipelines are in deal['pipeline_id'] as int)
    # We'll build this from the first few deals + deal_fields
    # For now, read all deals
    raw = json.loads(path.read_text())

    # Build pipeline_id → name from deals (Pipedrive includes pipeline info)
    pid_to_name = {}
    for deal in raw[:100]:
        pid = deal.get("pipeline_id")
        # pipeline name might be in a nested object or we need to infer
        if pid and pid not in pid_to_name:
            # Try to find from stage info or we'll map later
            pass

    # Map status: Pipedrive uses "open", "won", "lost"
    # Our data uses "진행 중", "성사됨", "실패"
    status_map = {"open": "진행 중", "won": "성사됨", "lost": "실패"}

    # Need pipeline names — fetch from Pipedrive API or use deal_fields
    # Since we have deal_fields with pipeline options, try that
    # Actually, pipeline_id maps to pipeline names via /pipelines endpoint
    # For now, let's check what pipeline_id values exist and map them
    if not pid_to_name:
        pid_counts = {}
        for deal in raw[:5000]:
            pid = deal.get("pipeline_id")
            if pid:
                pid_counts[pid] = pid_counts.get(pid, 0) + 1
        print(f"      pipeline_ids found: {pid_counts}")
        # We need to call API or use known mapping
        # Let's try to get pipeline names from the API
        import os, urllib.request, urllib.parse
        token = os.environ.get("PIPEDRIVE_API_TOKEN", "")
        domain = os.environ.get("PIPEDRIVE_DOMAIN", "")
        if token and domain:
            try:
                url = f"https://{domain}.pipedrive.com/api/v1/pipelines?api_token={token}"
                resp = urllib.request.urlopen(url, timeout=15)
                pdata = json.loads(resp.read())
                for p in (pdata.get("data") or []):
                    pid_to_name[p["id"]] = p["name"]
                print(f"      pipeline names: {pid_to_name}")
            except Exception as e:
                print(f"      pipeline fetch failed: {e}")

    claims = []
    for deal in raw:
        rec = {}
        for key, var_name in key_to_var.items():
            rec[var_name] = deal.get(key)
        # standard fields
        raw_status = deal.get("status", "")
        rec["status"] = status_map.get(raw_status, raw_status)
        pid = deal.get("pipeline_id")
        rec["pipeline"] = pid_to_name.get(pid, str(pid))
        claims.append(rec)
    print(f"      loaded {len(claims):,} deals from JSON")
    return claims


def _load_xlsx(path: Path) -> list[dict]:
    """xlsx 직접 로드 (offline fallback)."""
    import openpyxl
    print(f"  Loading XLSX: {path.name}")
    # 컬럼 인덱스 매핑 (이전 분석에서 확정)
    IDX = {
        0: "apply_date", 3: "status", 8: "payment_amount",
        9: "decision_amount", 11: "apply_amount", 18: "filing_date",
        45: "filing_amount", 101: "decision_date", 127: "payment_date",
        153: "pipeline", 62: "is_only_gam",
    }
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)  # header
    claims = []
    for row in it:
        rec = {}
        for idx, var in IDX.items():
            rec[var] = row[idx] if idx < len(row) else None
        claims.append(rec)
    wb.close()
    print(f"      loaded {len(claims):,} deals from XLSX")
    return claims


# ── aggregate to monthly series ──────────────────────
def aggregate(claims: list[dict]):
    """claims → 월별 집계 시리즈 (정기 B / 추심 C)."""
    series = {
        grp: {k: defaultdict(float) for k in ["app", "fil", "dec", "pay"]}
        for grp in ("B", "C")
    }
    for c in claims:
        status = str(c.get("status", ""))
        pipe = str(c.get("pipeline", ""))
        if STATUS_EXCLUDE in status:
            continue
        if PIPELINE_REGULAR in pipe:
            grp = "B"
        elif any(p in pipe for p in PIPELINE_COLLECTION):
            grp = "C"
        else:
            continue
        ad = parse_date(c.get("apply_date"))
        fd = parse_date(c.get("filing_date"))
        dd = parse_date(c.get("decision_date"))
        pd_ = parse_date(c.get("payment_date"))
        if ad: series[grp]["app"][ym(ad)] += to_num(c.get("apply_amount"))
        if fd: series[grp]["fil"][ym(fd)] += to_num(c.get("filing_amount"))
        if dd: series[grp]["dec"][ym(dd)] += to_num(c.get("decision_amount"))
        if pd_: series[grp]["pay"][ym(pd_)] += to_num(c.get("payment_amount"))
    return series


# ── cohort distribution fitting ──────────────────────
def fit_dist(claims, src_date_key, src_amt_key, tgt_date_key, tgt_amt_key,
             max_off, last_complete_m, window=None, min_amt=1e8, pipe_filter="B"):
    """Cohort distribution with optional rolling window."""
    src_total = defaultdict(float)
    matrix = defaultdict(lambda: defaultdict(float))
    for c in claims:
        status = str(c.get("status", ""))
        pipe = str(c.get("pipeline", ""))
        if STATUS_EXCLUDE in status:
            continue
        if pipe_filter == "B" and PIPELINE_REGULAR not in pipe:
            continue
        if pipe_filter == "C" and not any(p in pipe for p in PIPELINE_COLLECTION):
            continue
        sd = parse_date(c.get(src_date_key))
        td = parse_date(c.get(tgt_date_key))
        sa = to_num(c.get(src_amt_key))
        ta = to_num(c.get(tgt_amt_key))
        if sd and sa > 0:
            src_total[ym(sd)] += sa
        if sd and td and sa > 0 and ta > 0:
            off = ym(td) - ym(sd)
            if 0 <= off <= max_off:
                matrix[ym(sd)][off] += ta
    offs = defaultdict(list)
    valid_max = last_complete_m - max_off
    valid_min = (valid_max - window + 1) if window else None
    for src_m, row in matrix.items():
        if src_m > valid_max:
            continue
        if valid_min is not None and src_m < valid_min:
            continue
        sa = src_total[src_m]
        if sa < min_amt:
            continue
        for off in range(max_off + 1):
            offs[off].append(row.get(off, 0) / sa * 100)
    return {off: float(np.mean(v)) for off, v in offs.items() if v}


# ── prediction engine ────────────────────────────────
class ForecastEngine:
    def __init__(self, claims, series, current_partial_m, today_day=None):
        self.claims = claims
        self.series = series
        self.current = current_partial_m
        self.last_complete = current_partial_m - 1
        self.today_day = today_day or datetime.now().day

        # fit distributions
        lc = self.last_complete
        w = ROLLING_WINDOW
        self.a2f = fit_dist(claims, "apply_date", "apply_amount", "filing_date", "filing_amount",
                            CHAIN_DIST_MAX_OFF["a2f"], lc, window=w)
        self.f2d = fit_dist(claims, "filing_date", "filing_amount", "decision_date", "decision_amount",
                            CHAIN_DIST_MAX_OFF["f2d"], lc, window=w)
        self.d2p = fit_dist(claims, "decision_date", "decision_amount", "payment_date", "payment_amount",
                            CHAIN_DIST_MAX_OFF["d2p"], lc, window=w)

        # fallback: recent avg application
        app = series["B"]["app"]
        recent = [app.get(m, 0) for m in [lc - 2, lc - 1, lc]]
        self.app_avg = float(np.mean([v for v in recent if v > 0])) if any(v > 0 for v in recent) else 0

        # collection: 채권풀 기반 예측
        self._init_collection_pool(claims, series)

    def _init_collection_pool(self, claims, _series):
        """채권풀 잔액 × 월간 회수율(utilization rate) 기반 추심 예측."""
        lc = self.last_complete
        col_pipes = PIPELINE_COLLECTION

        # 추심 건 파싱: 신청월, 결정액, 결제월, 결제액
        col_deals = []
        for c in claims:
            pipe = str(c.get("pipeline", ""))
            status = str(c.get("status", ""))
            if STATUS_EXCLUDE in status:
                continue
            if not any(p in pipe for p in col_pipes):
                continue
            ad = parse_date(c.get("apply_date"))
            dd = parse_date(c.get("decision_date"))
            pd_ = parse_date(c.get("payment_date"))
            dec_amt = to_num(c.get("decision_amount"))
            pay_amt = to_num(c.get("payment_amount"))
            if not ad:
                continue
            col_deals.append({
                "apply_m": ym(ad),
                "dec_amt": dec_amt,
                "pay_m": ym(pd_) if pd_ and pay_amt > 0 else None,
                "pay_amt": pay_amt if pd_ else 0,
            })

        def pool_balance(T):
            return sum(d["dec_amt"] for d in col_deals
                       if d["apply_m"] < T and (d["pay_m"] is None or d["pay_m"] >= T))

        def actual_pay(T):
            return sum(d["pay_amt"] for d in col_deals if d["pay_m"] == T)

        # 최근 3개월 utilization rate (결제/풀잔액)
        rates = []
        for i in range(1, COLLECTION_MA_WINDOW + 1):
            T = lc - i + 1
            pool = pool_balance(T)
            paid = actual_pay(T)
            if pool > 0:
                rates.append(paid / pool)
        self.col_util_rate = float(np.mean(rates)) if rates else 0

        # 현재 풀 잔액
        self.col_pool = pool_balance(self.current)

        # 풀 순변동: 최근 3개월 실측 기반
        # (B→추심 유입을 간접 공식으로 추정하면 d2p에 수수료율이 내포되어
        #  미결제율을 8배 과대추정하는 문제가 있으므로, 실측 변동을 직접 사용)
        pools = [pool_balance(lc - i) for i in range(COLLECTION_MA_WINDOW, -1, -1)]
        deltas = [pools[i + 1] - pools[i] for i in range(len(pools) - 1)]
        self.col_pool_delta = float(np.mean(deltas)) if deltas else 0

        self._col_deals = col_deals
        self._col_pool_balance = pool_balance

    def _predict_collection(self, months_ahead):
        """추심 결제 예측: 실측 풀 순변동 기반."""
        pool_est = self.col_pool + self.col_pool_delta * months_ahead
        pool_est = max(pool_est, 0)
        return pool_est * self.col_util_rate

    def _backtest_collection(self, target_m):
        """백테스트용: target_m 시점의 추심 예측 (직전 3개월 rate 사용)."""
        def pool_balance(T):
            return sum(d["dec_amt"] for d in self._col_deals
                       if d["apply_m"] < T and (d["pay_m"] is None or d["pay_m"] >= T))
        def actual_pay(T):
            return sum(d["pay_amt"] for d in self._col_deals if d["pay_m"] == T)

        pool = pool_balance(target_m)
        rates = []
        for j in range(1, COLLECTION_MA_WINDOW + 1):
            T = target_m - j
            p = pool_balance(T)
            a = actual_pay(T)
            if p > 0:
                rates.append(a / p)
        rate = float(np.mean(rates)) if rates else self.col_util_rate
        return pool * rate

    def _get_app(self, m):
        s = self.series["B"]["app"]
        if m <= self.last_complete and m in s:
            return s[m]
        if m == self.current and m in s:
            return s[m] * 30 / self.today_day
        return self.app_avg

    def _get_fil(self, m):
        s = self.series["B"]["fil"]
        if m <= self.last_complete and m in s:
            return s[m]
        if m == self.current and m in s:
            return s[m] * 30 / self.today_day
        return sum(self._get_app(m - off) * self.a2f.get(off, 0) / 100 for off in self.a2f)

    def _get_dec(self, m):
        s = self.series["B"]["dec"]
        if m <= self.last_complete and m in s:
            return s[m]
        if m == self.current and m in s:
            return s[m] * 30 / self.today_day
        return sum(self._get_fil(m - off) * self.f2d.get(off, 0) / 100 for off in self.f2d)

    def predict(self, target_m) -> dict:
        """단일 월 예측."""
        pred_reg = 0
        breakdown = []
        for off, r in self.d2p.items():
            src_m = target_m - off
            d = self._get_dec(src_m)
            contrib = d * r / 100
            pred_reg += contrib
            is_actual = src_m <= self.last_complete and src_m in self.series["B"]["dec"]
            breakdown.append({
                "off": off, "src": ym_label(src_m),
                "dec_amount": round(d / 1e8, 2), "rate": r,
                "contrib": round(contrib / 1e8, 2),
                "source": "실측" if is_actual else "추정"
            })
        months_ahead = target_m - self.current
        col = self._predict_collection(months_ahead)
        pred_total = pred_reg + col
        mon = month_of(target_m)
        adj = SEASON_ADJUSTMENT.get(mon, 0)
        adjusted = pred_total * (1 + adj)
        return {
            "month": ym_label(target_m),
            "regular": round(pred_reg / 1e8, 2),
            "collection": round(col / 1e8, 2),
            "total": round(pred_total / 1e8, 2),
            "season_adj": adj,
            "adjusted": round(adjusted / 1e8, 2),
            "breakdown": breakdown,
        }

    def forecast(self, n_months=5) -> list[dict]:
        """현재월부터 n_months개 예측."""
        return [self.predict(self.current + i) for i in range(n_months)]

    def backtest(self, n_months=12) -> list[dict]:
        """과거 n_months Walk-Forward 검증."""
        results = []
        for i in range(n_months, 0, -1):
            tgt = self.current - i
            actual_b = self.series["B"]["pay"].get(tgt, 0) / 1e8
            actual_c = self.series["C"]["pay"].get(tgt, 0) / 1e8
            actual = actual_b + actual_c

            # refit at that time
            lc = tgt - 1
            w = ROLLING_WINDOW
            d2p_t = fit_dist(self.claims, "decision_date", "decision_amount",
                             "payment_date", "payment_amount",
                             CHAIN_DIST_MAX_OFF["d2p"], lc, window=w)
            f2d_t = fit_dist(self.claims, "filing_date", "filing_amount",
                             "decision_date", "decision_amount",
                             CHAIN_DIST_MAX_OFF["f2d"], lc, window=w)
            # simple prediction using actual series
            pred_b = 0
            bdec = self.series["B"]["dec"]
            bfil = self.series["B"]["fil"]
            for off, r in d2p_t.items():
                sm = tgt - off
                d = bdec.get(sm, 0)
                if d == 0 and sm > lc:
                    d = sum(bfil.get(sm - o2, 0) * f2d_t.get(o2, 0) / 100 for o2 in f2d_t)
                pred_b += d * r / 100

            pred_c = self._backtest_collection(tgt)
            pred = (pred_b + pred_c) / 1e8
            mon = month_of(tgt)
            # forecast와 동일한 시즌 보정 적용 (미적용 시 March -15.3% 등 과대 오차)
            adj = SEASON_ADJUSTMENT.get(mon, 0)
            pred_adj = pred * (1 + adj)
            err = (pred_adj - actual) / actual * 100 if actual > 0 else 0
            results.append({
                "month": ym_label(tgt),
                "actual": round(actual, 2),
                "predicted": round(pred_adj, 2),
                "error_pct": round(err, 1),
                "season_month": mon,
                "is_season_outlier": abs(err) > 15,
            })
        return results


# ── 법인 간이 모델 ─────────────────────────────────────
def load_corp_deals() -> list[dict]:
    """법인 slim JSON 로드."""
    path = DATA_DIR / "deals_corp_slim.json"
    if not path.exists():
        return []
    print(f"  Loading corp JSON: {path.name} ({path.stat().st_size / 1e6:.1f} MB)")
    raw = json.loads(path.read_text())
    status_map = {"open": "진행 중", "won": "성사됨", "lost": "실패"}
    for d in raw:
        s = d.get("status", "")
        d["status"] = status_map.get(s, s)
    print(f"      loaded {len(raw):,} corp deals")
    return raw


def aggregate_corp(claims):
    """법인 claims → 월별 결제 시계열 (정기/추심)."""
    pay = {"regular": defaultdict(float), "collection": defaultdict(float)}
    for c in claims:
        status = str(c.get("status", ""))
        pipe = str(c.get("pipeline", ""))
        if STATUS_EXCLUDE in status:
            continue
        pd_ = parse_date(c.get("payment_date"))
        pa = to_num(c.get("payment_amount"))
        if not pd_ or pa <= 0:
            continue
        m = ym(pd_)
        if CORP_PIPELINE_REGULAR in pipe:
            pay["regular"][m] += pa
        elif any(p in pipe for p in CORP_PIPELINE_COLLECTION):
            pay["collection"][m] += pa
    return pay


def aggregate_corp_full(claims):
    """법인 claims → 월별 app/fil/dec/pay 시계열 (정기/추심)."""
    series = {
        grp: {k: defaultdict(float) for k in ["app", "fil", "dec", "pay"]}
        for grp in ("regular", "collection")
    }
    for c in claims:
        status = str(c.get("status", ""))
        pipe = str(c.get("pipeline", ""))
        if STATUS_EXCLUDE in status:
            continue
        if CORP_PIPELINE_REGULAR in pipe:
            grp = "regular"
        elif any(p in pipe for p in CORP_PIPELINE_COLLECTION):
            grp = "collection"
        else:
            continue
        ad = parse_date(c.get("apply_date"))
        fd = parse_date(c.get("filing_date"))
        dd = parse_date(c.get("decision_date"))
        pd_ = parse_date(c.get("payment_date"))
        if ad: series[grp]["app"][ym(ad)] += to_num(c.get("apply_amount"))
        if fd: series[grp]["fil"][ym(fd)] += to_num(c.get("filing_amount"))
        if dd: series[grp]["dec"][ym(dd)] += to_num(c.get("decision_amount"))
        if pd_: series[grp]["pay"][ym(pd_)] += to_num(c.get("payment_amount"))
    return series


def series_to_list(s, current_m, n=24):
    """월별 defaultdict → [{month, amount(억)}, ...] (current_m 포함 최근 n개월)."""
    return [
        {"month": ym_label(m), "amount": round(s.get(m, 0) / 1e8, 3)}
        for m in range(current_m - n + 1, current_m + 1)
    ]


class CorpForecastEngine:
    """법인 간이 예측: 6개월 이동평균 + 선형추세."""

    def __init__(self, corp_pay, current_m):
        self.pay = corp_pay
        self.current = current_m
        self.last_complete = current_m - 1
        lc = self.last_complete

        # 정기/추심 각각 최근 6개월 MA + 추세
        self.reg_params = self._fit_trend(corp_pay["regular"], lc)
        self.col_params = self._fit_trend(corp_pay["collection"], lc)

    def _fit_trend(self, series, lc, window=6):
        """최근 window개월 데이터로 선형 추세 계산."""
        vals = []
        for i in range(window):
            m = lc - window + 1 + i
            vals.append(series.get(m, 0) / 1e8)
        x = np.arange(len(vals))
        if all(v == 0 for v in vals):
            return {"ma": 0, "slope": 0}
        slope, intercept = np.polyfit(x, vals, 1)
        ma = float(np.mean(vals))
        return {"ma": ma, "slope": float(slope), "last": float(vals[-1]),
                "intercept": float(intercept), "window": len(vals)}

    def _predict_series(self, params, months_ahead):
        """선형추세로 예측 (억 단위)."""
        idx = params["window"] - 1 + months_ahead
        pred = params["intercept"] + params["slope"] * idx
        return max(pred, 0)

    def forecast(self, n_months=5):
        results = []
        for i in range(n_months):
            target_m = self.current + i
            reg = self._predict_series(self.reg_params, i)
            col = self._predict_series(self.col_params, i)
            total = reg + col
            results.append({
                "month": ym_label(target_m),
                "regular": round(reg, 3),
                "collection": round(col, 3),
                "total": round(total, 3),
            })
        return results

    def backtest(self, n_months=12):
        results = []
        for i in range(n_months, 0, -1):
            tgt = self.current - i
            actual_r = self.pay["regular"].get(tgt, 0) / 1e8
            actual_c = self.pay["collection"].get(tgt, 0) / 1e8
            actual = actual_r + actual_c
            # walk-forward: 직전 6개월 MA
            pred_vals = []
            for j in range(1, 7):
                m = tgt - j
                pred_vals.append((self.pay["regular"].get(m, 0) +
                                  self.pay["collection"].get(m, 0)) / 1e8)
            pred = float(np.mean(pred_vals)) if pred_vals else 0
            err = (pred - actual) / actual * 100 if actual > 0 else 0
            results.append({
                "month": ym_label(tgt),
                "actual": round(actual, 3),
                "predicted": round(pred, 3),
                "error_pct": round(err, 1),
            })
        return results


# ── main ─────────────────────────────────────────────
def main():
    print("=" * 60)
    print("비즈넵 결제 예측 모델 — Phase 2 v2 코호트 분산")
    print("=" * 60)

    claims = load_deals()
    print(f"  Total claims: {len(claims):,}")

    series = aggregate(claims)
    # Determine current partial month
    all_ms = set()
    for grp in series.values():
        for s in grp.values():
            all_ms.update(s.keys())
    current_m = max(all_ms)
    print(f"  Data range: ... ~ {ym_label(current_m)}")

    engine = ForecastEngine(claims, series, current_m)

    # Distributions
    dists = {
        "a2f": {str(k): round(v, 2) for k, v in engine.a2f.items()},
        "f2d": {str(k): round(v, 2) for k, v in engine.f2d.items()},
        "d2p": {str(k): round(v, 2) for k, v in engine.d2p.items()},
    }
    print(f"\n[분산 비율]")
    for name, d in dists.items():
        print(f"  {name}: {d}")

    # Collection pool info
    print(f"\n[추심 채권풀]")
    print(f"  풀 잔액: {engine.col_pool / 1e8:.1f}억")
    print(f"  월간 회수율: {engine.col_util_rate * 100:.3f}%")
    print(f"  풀 순변동: 월 {engine.col_pool_delta / 1e8:+.1f}억")

    # Forecast
    print(f"\n[향후 5개월 예측]")
    fc = engine.forecast(5)
    for f in fc:
        print(f"  {f['month']}: 통합 {f['total']}억 → 시즌보정 {f['adjusted']}억 (adj {f['season_adj']:+.0%})")

    # Backtest
    print(f"\n[12개월 백테스트]")
    bt = engine.backtest(12)
    mape = float(np.mean([abs(r["error_pct"]) for r in bt]))
    for r in bt:
        tag = "★" if abs(r["error_pct"]) <= 10 else ("⚠" if abs(r["error_pct"]) <= 20 else "✗")
        print(f"  {r['month']}: 실제 {r['actual']}억 vs 예측 {r['predicted']}억 ({r['error_pct']:+.1f}%) {tag}")
    print(f"  MAPE: {mape:.1f}%")

    # ── 법인 ──
    corp_claims = load_corp_deals()
    corp_fc = []
    corp_bt = []
    corp_mape = None
    if corp_claims:
        corp_pay = aggregate_corp(corp_claims)
        corp_engine = CorpForecastEngine(corp_pay, current_m)

        print(f"\n[법인 예측]")
        print(f"  정기 MA: {corp_engine.reg_params['ma']:.3f}억 (slope {corp_engine.reg_params['slope']:+.4f})")
        print(f"  추심 MA: {corp_engine.col_params['ma']:.3f}억 (slope {corp_engine.col_params['slope']:+.4f})")

        corp_fc = corp_engine.forecast(5)
        for f in corp_fc:
            print(f"  {f['month']}: 정기 {f['regular']}억 + 추심 {f['collection']}억 = {f['total']}억")

        corp_bt = corp_engine.backtest(12)
        corp_mape = float(np.mean([abs(r["error_pct"]) for r in corp_bt if r["actual"] > 0]))
        print(f"  법인 MAPE: {corp_mape:.1f}%")

    # ── 통합 ──
    print(f"\n[통합 예측 (개인+법인)]")
    combined_fc = []
    for i, f_ind in enumerate(fc):
        f_corp = corp_fc[i] if i < len(corp_fc) else {"regular": 0, "collection": 0, "total": 0}
        combined = {
            "month": f_ind["month"],
            "individual": {"regular": f_ind["regular"], "collection": f_ind["collection"],
                           "total": f_ind["total"], "season_adj": f_ind["season_adj"],
                           "adjusted": f_ind["adjusted"], "breakdown": f_ind["breakdown"]},
            "corporate": f_corp,
            "grand_total": round(f_ind["adjusted"] + f_corp["total"], 2),
        }
        combined_fc.append(combined)
        print(f"  {combined['month']}: 개인 {f_ind['adjusted']}억 + 법인 {f_corp['total']}억 = {combined['grand_total']}억")

    # 월별 시리즈 (app/fil/dec/pay × 개인 B/C + 법인 regular/collection, 최근 24개월)
    corp_series = aggregate_corp_full(corp_claims) if corp_claims else {
        "regular": {k: defaultdict(float) for k in ["app", "fil", "dec", "pay"]},
        "collection": {k: defaultdict(float) for k in ["app", "fil", "dec", "pay"]},
    }
    monthly_series = {
        "individual": {
            grp_name: {stage: series_to_list(series[grp][stage], current_m)
                       for stage in ["app", "fil", "dec", "pay"]}
            for grp, grp_name in [("B", "regular"), ("C", "collection")]
        },
        "corporate": {
            grp: {stage: series_to_list(corp_series[grp][stage], current_m)
                  for stage in ["app", "fil", "dec", "pay"]}
            for grp in ["regular", "collection"]
        },
    }

    # Save outputs
    output = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "data_range": f"... ~ {ym_label(current_m)}",
        "total_claims": len(claims),
        "total_corp_claims": len(corp_claims),
        "distributions": dists,
        "collection_pool": {
            "balance": round(engine.col_pool / 1e8, 1),
            "utilization_rate": round(engine.col_util_rate * 100, 3),
            "monthly_delta": round(engine.col_pool_delta / 1e8, 1),
        },
        "season_adjustments": SEASON_ADJUSTMENT,
        "monthly_series": monthly_series,
        "forecast": combined_fc,
        "backtest": bt,
        "corp_backtest": corp_bt,
        "mape": round(mape, 2),
        "corp_mape": round(corp_mape, 1) if corp_mape else None,
    }
    out_path = OUTPUT_DIR / "forecast.json"
    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2))
    print(f"\n→ {out_path}")


if __name__ == "__main__":
    main()
